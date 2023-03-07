def converandrename ():
    
    '''
COPYRIGHTS TO DR.JOSEPH NADY FOUAD
'''
from tkinter import *
from tkinter import filedialog
import os
import time

from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from win32com.client import Dispatch
from datetime import datetime
from datetime import date
from openpyxl import Workbook
import win32com.client as client
from openpyxl import load_workbook
import selenium.common.exceptions

speak = Dispatch("SAPI.SpVoice").Speak

#GUI____________________________________________________
window2= Tk()
window2.geometry('300x300+800+200')
window2.resizable(False, False)
window2.title('JN CRM Solutions')
window2.configure(background="silver")


#Functions
def importdir ():
    filename = filedialog.askdirectory()
    importdirEntry.insert(0,filename) 

def exportdir ():
    filename = filedialog.askdirectory()
    exportdirEntry.insert(0,filename) 

def ConvertAndRenameSheets():
    window2.quit()

#Labels
title = Label(window2, text="File Converter ", bg='white', fg='red',font=("Arial",20))
title.pack(fill=X)

importdirlabel = Label(window2, text="Choose folder contains the files to be converted:", bg='silver', fg='black',font=("Arial",8))
importdirlabel.place(x=10,y=80)

exportdirlabel = Label(window2, text="Choose Export Directory: ", bg='silver', fg='black',font=("Arial",8))
exportdirlabel.place(x=10,y=130)

converttolabel = Label(window2, text="Convert to:", bg='silver', fg='black',font=("Arial",8))
converttolabel.place(x=10,y=180)

#drop list
variable = StringVar(window2)
variable.set("xlsx") # default value

w = OptionMenu(window2, variable, "xlsx", "csv","csv utf-8")
w.place(x=10,y=200)

#entry
importdirEntry = Entry(window2,width=20)
importdirEntry.place(x=10, y=100)

exportdirEntry = Entry(window2,width=20)
exportdirEntry.place(x=10, y=150)

#button
importbtn = Button(window2, text = "Browse", justify='center', width=10, height=1, command= importdir)
importbtn.place(x=200, y=100)

Exportbtn = Button(window2, text = "Browse", justify='center', width=10, height=1, command= exportdir)
Exportbtn.place(x=200, y=150)

quitbtn = Button(window2, text = "Convert", justify='center', width=10, height=1, command= ConvertAndRenameSheets)
quitbtn.place(x=100, y=250)

# _____________________Footer_______________________
footer = Label(window2, text="Developed by Dr. Joseph Nady",
         bg="silver",fg='black', font=("Arial", 7))
footer.place(x=70, y=280)

window2.mainloop()



#__________________________________________________________________________________________________
#SCRIPT
#__________________________________________________________________________________________________
importDir = importdirEntry.get()
exportDir =  exportdirEntry.get()
Convertto = variable.get()
type(Convertto)
print(f"import directory:\n{importDir}\nexport directory:\n {exportDir}\nextension:\n {Convertto}")#test

excel = client.Dispatch("Excel.application")

def convert():
    for file in os.listdir(importDir):
        filename, fileextension = os.path.splitext(file)
        print(f">>>>   START CONVERSION OF {filename+fileextension}   <<<<")
        import_path = f"{importDir}/{file}"
        wb = excel.Workbooks.Open(import_path)
        
        #os.path.normpath (normalize the path from string to path C:/User/ >> C://User//)
        output_path= os.path.normpath(f"{exportDir}/{filename}")

        #6 >> CSV -- 51 >> xlsx -- 62 >> CSV UTF-8
        if Convertto == "xlsx":
            wb.SaveAs(output_path,51)
        elif Convertto == "csv":
            wb.SaveAs(output_path,6)
        elif Convertto == "csv utf-8":
            wb.SaveAs(output_path,62)

        wb.Close(True)

    speak("FILES CONVERTED SUCCESSFULLY")
    print(">>>>   FILES CONVERTED SUCCESSFULLY   <<<<")

def sheet_rename(title):
    for file in os.listdir(exportDir):
        wbpath = f"{exportDir}/{file}"
        workbook = load_workbook(wbpath)
        workbook.sheetnames
        sheet = workbook.active
        print(f">>>>   START RENAMING OF {file}   <<<<")
        sheet.title = title
        #os.path.normpath (normalize the path from string to path C:/User/ >> C://User//)
        workbook.save(os.path.normpath(f"{exportDir}/{file}"))
    speak("FILES RENAMED SUCCESSFULLY")
    print(">>>>   FILES RENAMED SUCCESSFULLY   <<<<")

convert()
sheet_rename('Sheet1')
excel.Quit()